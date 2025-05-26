import os
import numpy as np
import torch
import torch.nn as nn
from model import *
from function import *
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.metrics import mean_squared_error
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

torch.manual_seed(123)
torch.cuda.manual_seed_all(123)


def to_gpu(x):
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    return x.to(device)

class Solver(object):
    def __init__(self, config,
                train_dataloaders,valid_dataloaders,test_dataloaders,
                finetune_train_dataloaders,finetune_test_dataloaders,
                increase_train_dataloaders,increase_test_dataloaders,
                stim_dataloaders,
                stim_list,label_process,
                finetune_id=0, model=None):

        self.config = config
        self.train_dataloaders = train_dataloaders
        self.valid_dataloaders = valid_dataloaders
        self.test_dataloaders = test_dataloaders

        self.finetune_train_dataloaders = finetune_train_dataloaders
        self.finetune_test_dataloaders = finetune_test_dataloaders

        self.increase_train_dataloaders = increase_train_dataloaders
        self.increase_test_dataloaders = increase_test_dataloaders

        self.stim_dataloaders = stim_dataloaders
        self.label_process = label_process
        self.model = model
        self.finetune_id = finetune_id

        self.stim_mse, self.stim_mae, self.stim_corr, self.stim_nrmse = None, None, None, None

        self.train_losses, self.train_maes, self.train_corrs, self.train_nrmses = [], [], [], []
        self.valid_losses, self.valid_maes, self.valid_corrs, self.valid_nrmses = [], [], [], []

        self.finetune_before_label = []
        self.finetune_before_pred = []
        self.finetune_label = []
        self.finetune_pred = []
        
        self.valid_label=[]
        self.valid_pred=[]
        self.before_label=[]
        self.before_pred=[]
        self.after_label=[]
        self.after_pred=[]
        self.stim_before_label = []
        self.stim_before_pred = []
        self.stim_after_label = []
        self.stim_after_pred = []

        self.stim_list = stim_list

        self.finetune_before_mse, self.finetune_before_mae, self.finetune_before_corr, self.finetune_before_nrmse = None, None, None, None
        self.finetune_after_mse, self.finetune_after_mae, self.finetune_after_corr, self.finetune_after_nrmse = None, None, None, None
        self.increase_before_mse, self.increase_before_mae, self.increase_before_corr, self.increase_before_nrmse = None, None, None, None
        self.increase_after_mse, self.increase_after_mae, self.increase_after_corr, self.increase_after_nrmse = None, None, None, None
        self.stim_before_mse, self.stim_before_mae, self.stim_before_corr, self.stim_before_nrmse = None, None, None, None
        self.stim_after_mse, self.stim_after_mae, self.stim_after_corr, self.stim_after_nrmse = None, None, None, None

    def draw(self,y_true, y_pred, ps=""):

        plt.plot(y_true, label='True Values', color="blue")
        plt.plot(y_pred, label='Predictions', color="red")
        plt.xlabel('Index')
        plt.ylabel('Value')
        plt.legend()
        plt.savefig(self.config.result_path+"/"+"Experiment"+self.config.time+ps+"_id_"+str(self.config.finetune_id)+".jpg")
        plt.show()

    def build(self, cuda=True):

        if self.config.model == "Model_LSTM":
            self.model = Model_LSTM()   

        to_gpu(self.model)

        print(torch.cuda.get_device_capability())

        torch.backends.cudnn.benchmark = True

        self.optimizer = self.config.optimizer(
            filter(lambda p: p.requires_grad, self.model.parameters()),
            lr=self.config.learning_rate,
            weight_decay=self.config.weight_decay)

    def train(self):
        curr_patience = patience = self.config.patience
        num_trials = 1

        self.criterion = criterion = nn.MSELoss(reduction="mean")


        best_valid_loss = float('inf')

        lr_scheduler = torch.optim.lr_scheduler.ExponentialLR(self.optimizer, gamma=0.5)


        self.model.train()
        print()
        for e in range(self.config.n_epoch):
            train_loss = []
            train_corr = []
            train_mae = []
            train_nrmse = []

            cnt = 0
            train_acc = 0
            train_cnt = 0
            cnt = cnt + 1
            print(f'--here--{cnt}')

            for batch in self.train_dataloaders: 
                self.model.zero_grad()

                x, y = batch

                
                y = self.label_process.standardize(y)
                x = to_gpu(x)
                y = to_gpu(y)

                self.model.train()
                y_pred = self.model(x)

                loss = self.criterion(y_pred, y)

                loss.backward()
                self.optimizer.step()

                y_np = y.detach().cpu().numpy()
                y_pred_np = y_pred.detach().cpu().numpy()

                y_np = self.label_process.inverse_standardize(y_np)
                y_pred_np = self.label_process.inverse_standardize(y_pred_np)

                mse, mae, corr, nrmse = caculate(y_pred_np, y_np)
                
                train_loss.append(mse)
                train_mae.append(mae)
                train_corr.append(corr)
                train_nrmse.append(nrmse)

            valid_loss, valid_mae, valid_corr, valid_nrmse = self.eval(data="valid")



            self.train_losses.append(np.nanmean((train_loss)))
            self.train_corrs.append(np.nanmean(train_corr))
            self.train_maes.append(np.nanmean(train_mae))
            self.train_nrmses.append(np.nanmean(train_nrmse))

            self.valid_losses.append(valid_loss)
            self.valid_corrs.append(valid_corr)
            self.valid_maes.append(valid_mae)
            self.valid_nrmses.append(valid_nrmse)



            print(f"Epoch {e + 1}/{self.config.n_epoch}, train loss: {np.mean(train_loss)}, valid loss: {valid_loss}")
            print(f"Valid mae: {valid_mae}, valid corr: {valid_corr}, valid nrmse: {valid_nrmse}")\

            print(f"Current patience: {curr_patience}, current trial: {num_trials}.")
            if valid_loss <= best_valid_loss:
                best_valid_loss = valid_loss
                print("Found new best model on dev set!")
                if not os.path.exists('checkpoints'):
                    os.makedirs('checkpoints')
                torch.save(self.model.state_dict(), f'checkpoints/model_{self.finetune_id}_{self.config.name}.std')
                torch.save(self.optimizer.state_dict(), f'checkpoints/optim_{self.config.name}.std')
                curr_patience = patience
            else:
                curr_patience -= 1
                if curr_patience <= -1:
                    print("Running out of patience, loading previous best model.")
                    num_trials -= 1
                    curr_patience = patience
                    self.model.load_state_dict(torch.load(f'checkpoints/model_{self.finetune_id}_{self.config.name}.std'))
                    self.optimizer.load_state_dict(torch.load(f'checkpoints/optim_{self.config.name}.std'))
                    lr_scheduler.step()
                    print(f"Current learning rate: {self.optimizer.state_dict()['param_groups'][0]['lr']}")

            if num_trials <= 0:
                print("Running out of patience, early stopping.")
                break


        self.model.load_state_dict(torch.load(f'checkpoints/model_{self.finetune_id}_{self.config.name}.std'))
        self.valid_mse, self.valid_mae, self.valid_corr, self.valid_nrmse = self.eval(data="valid",record=True)
        self.test_mse, self.test_mae, self.test_corr, self.test_nrmse = self.eval(data="test",record=True)

    def eval(self,data="valid",mode=None,draw_pred=False,freeze=False,record=False):
        self.model.eval()
        true=[]
        pred=[]
        if data=="valid":
            dataloaders = self.valid_dataloaders
        elif data=="test":
            dataloaders = self.test_dataloaders
        elif data=="increase_valid":
            dataloaders = self.increase_valid_dataloaders
        elif data=="increase_test":
            dataloaders = self.increase_test_dataloaders
        elif data=="finetune_valid":
            dataloaders = self.finetune_valid_dataloaders
        elif data=="finetune_test":
            dataloaders = self.finetune_test_dataloaders
            
        with torch.no_grad():
            for batch in dataloaders:
                x, y = batch
                x = to_gpu(x)
                y = to_gpu(y)

                y = self.label_process.standardize(y)

                y_pred = self.model(x)

                y_np = self.label_process.inverse_standardize(y)
                y_np_pred = self.label_process.inverse_standardize(y_pred)

                for i in y_np.detach().cpu().numpy().flatten():
                    true.append(i)
                for i in y_np_pred.detach().cpu().numpy().flatten():
                    pred.append(i)

        true = np.array(true).flatten()
        pred = np.array(pred).flatten()

        if record==True:
            if data=="valid":
                self.valid_label = true
                self.valid_pred = pred
            elif data=="test":
                self.test_label = true
                self.test_pred = pred
            elif data=="increase_valid":
                self.increase_valid_label = true
                self.increase_valid_pred = pred
            elif data=="increase_test":
                if mode=="before":
                    self.increase_before_label = true
                    self.increase_before_pred = pred
                elif mode=="after":
                    self.increase_after_label = true
                    self.increase_after_pred = pred
            elif data=="finetune_valid":
                self.finetune_valid_label = true
                self.finetune_valid_pred = pred
            elif data=="finetune_test":
                if mode=="before":
                    self.finetune_before_label = true
                    self.finetune_before_pred = pred
                elif mode=="after":
                    self.finetune_after_label = true
                    self.finetune_after_pred = pred

        mse,mae, corr, nrmse = caculate(pred, true)
        return mse, mae, corr, nrmse
    
    def finetune_train(self,data="finetune",draw_pred=False,freeze=True):
        self.model.load_state_dict(torch.load(f'checkpoints/model_{self.finetune_id}_{self.config.name}.std'))
        self.optimizer.load_state_dict(torch.load(f'checkpoints/optim_{self.config.name}.std'))
        

        for param_group in self.optimizer.param_groups:
            param_group['lr'] = self.config.learning_rate

        if data=="finetune":
            self.finetune_before_mse, self.finetune_before_mae,self.finetune_before_corr, self.finetune_before_nrmse = self.eval(data="finetune_test",mode="before",record=True)
        elif data=="increase":
            self.increase_before_mse, self.increase_before_mae,self.increase_before_corr, self.increase_before_nrmse = self.eval(data="increase_test",mode="before",record=True)
        
        if freeze:
            for i, child in enumerate(self.model.children()):
                if i < 2:
                    for param in child.parameters():
                        param.requires_grad = False
        true=[]
        pred=[]
        if data=="finetune":
            dataloaders = self.finetune_train_dataloaders
        elif data=="increase":
            dataloaders = self.increase_train_dataloaders
        
        best_valid_loss = float('inf')

        for e in range(15):
            print(f'--Finetune--{data}--{e+1}')
            self.model.train()
            for batch in dataloaders: 
                self.model.zero_grad()
                x, y = batch

                y = self.label_process.standardize(y)
                x = to_gpu(x)
                y = to_gpu(y)

                self.model.train()
                y_pred = self.model(x)

                loss = self.criterion(y_pred, y)
                loss.backward()
                self.optimizer.step()

                y_np = y.detach().cpu().numpy()
                y_pred_np = y_pred.detach().cpu().numpy()

                y_np = self.label_process.inverse_standardize(y_np)
                y_pred_np = self.label_process.inverse_standardize(y_pred_np)
                

                mse, mae, corr, nrmse = caculate(y_pred_np, y_np)


            if mse <= best_valid_loss:
                best_valid_loss = mse
                torch.save(self.model.state_dict(), f'checkpoints/model_{self.finetune_id}_{data}_{self.config.name}.std')
                torch.save(self.optimizer.state_dict(), f'checkpoints/optim_{data}_{self.config.name}.std')

        if data=="finetune":
            self.finetune_after_mse, self.finetune_after_mae, self.finetune_after_corr, self.finetune_after_nrmse = self.eval(data="finetune_test",mode="after",record=True)
        elif data=="increase":
            self.increase_after_mse, self.increase_after_mae, self.increase_after_corr, self.increase_after_nrmse = self.eval(data="increase_test",mode="after",record=True)

        if data=="increase":
            data = {
                    "Time": [self.config.time],
                    "Finetune_id": [self.config.finetune_id],
                    "Model": [self.config.model],
                    "Learning Rate" : [self.config.learning_rate],
                    'Valid MSE' : [self.valid_mse],
                    'Test MSE' : [self.test_mse],
                    'Increase_Before_MSE': [self.increase_before_mse],
                    'Increase_After_MSE': [self.increase_after_mse],
                    'Valid MAE' : [self.valid_mae],
                    'Test MAE' : [self.test_mae],
                    'Increase_Before_MAE': [self.increase_before_mae],
                    'Increase_After_MAE': [self.increase_after_mae],
                    'Valid NRMSE' : [self.valid_nrmse],
                    'Test NRMSE' : [self.test_nrmse],
                    'Increase_Before_NRMSE': [self.increase_before_nrmse],
                    'Increase_After_NRMSE': [self.increase_after_nrmse]
                }
            df = pd.DataFrame(data)

            file_path = self.config.result_path + "/" + 'Increase_results.xlsx'
            if os.path.exists(file_path):
                df_before = pd.read_excel(file_path)
                df_before = pd.DataFrame(df_before)
                df = pd.concat([df_before, df], ignore_index=True)
            df.to_excel(file_path, index=False)
            wb = Workbook()
            ws = wb.active
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            wb.save(file_path)

    
    def stim_test(self,mode='before',record=True):

        true=[]
        pred=[]
        if self.config.finetune_id in [18,20]:
            self.stim_before_label = true
            self.stim_before_pred = pred
            self.stim_after_label = true
            self.stim_after_pred = pred
            return None,None,None,None
            

        if mode =='before':
            self.model.load_state_dict(torch.load(f'checkpoints/model_{self.finetune_id}_finetune_{self.config.name}.std'))
        else:
            self.model.load_state_dict(torch.load(f'checkpoints/model_{self.finetune_id}_{self.config.name}.std'))
            
        with torch.no_grad():
            cnt = 0
            stim_file_result=[]
            for stim_test in self.stim_dataloaders:
                file_pred=[]
                file_true=[]
                for batch in stim_test:
                    self.model.zero_grad()
                    x, y = batch
                    x = to_gpu(x)
                    y = to_gpu(y)

                    y = self.label_process.standardize(y)

                    y_pred = self.model(x)
                    loss = self.criterion(y_pred, y)

                    y_np = y.detach().cpu().numpy()
                    y_pred_np = y_pred.detach().cpu().numpy()

                    y_np = self.label_process.inverse_standardize(y_np)
                    y_pred_np = self.label_process.inverse_standardize(y_pred_np)


                    for i in y_np.flatten():
                        true.append(i)
                        file_true.append(i)
                    for i in y_pred_np.flatten():
                        pred.append(i)
                        file_pred.append(i)

                file_pred = np.array(file_pred).flatten()
                file_true = np.array(file_true).flatten()
                temp_mse,temp_mae, temp_corr, temp_nrmse = caculate(file_pred, file_true)
                temp_string = str(self.stim_list[cnt])+" mse: "+str(temp_mse)+" mae:"+str(temp_mae)+" corr: "+str(temp_corr)+" nrmse: "+str(temp_nrmse)
                stim_file_result.append(temp_string)
                data = {
                    "Time": [self.config.time],
                    "Finetune_id": [self.config.finetune_id],
                    "Finetune_test_file": [str(self.stim_list[cnt])],
                    "Model": [self.config.model],
                    "Learning Rate" : [self.config.learning_rate],
                    'Before MSE' : [self.finetune_before_mse],
                    'Stim MSE' : [temp_mse],
                    'Before MAE' : [self.finetune_before_mae],
                    'Stim MAE' : [temp_mae],
                    'Before NRMSE' : [self.finetune_before_nrmse],
                    'Stim NRMSE' : [temp_nrmse],
                    "Ave Pred" : [np.array(file_pred).mean()],
                    "Ave IMS" : [np.array(file_true).mean()]
                }
                cnt+=1
                df = pd.DataFrame(data)

                file_path = self.config.result_path + "/" + 'Stim_results.xlsx'
                if os.path.exists(file_path):
                    df_before = pd.read_excel(file_path)
                    df_before = pd.DataFrame(df_before)
                    df = pd.concat([df_before, df], ignore_index=True)
                df.to_excel(file_path, index=False)
                wb = Workbook()
                ws = wb.active
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width

                wb.save(file_path)

        true = np.array(true).flatten()
        pred = np.array(pred).flatten()


        if mode == "before":
            self.stim_before_label = true
            self.stim_before_pred = pred
        elif mode == "after":
            self.stim_after_label = true
            self.stim_after_pred = pred

        mse, mae, corr, nrmse = caculate(pred, true)
        self.stim_mse, self.stim_mae, self.stim_corr, self.stim_nrmse = mse, mae, corr, nrmse
        return mse, mae, corr, nrmse


    def save_excel(self):
        data = {
            "Time": [self.config.time],
            "Finetune_id": [self.config.finetune_id],
            "Model": [self.config.model],
            "Learning Rate": [self.config.learning_rate],
            "Valid MSE": [self.valid_mse],
            "Test MSE": [self.test_mse],
            'Before MSE': [self.finetune_before_mse],
            'After MSE': [self.finetune_after_mse],
            "Valid MAE": [self.valid_mae],
            "Test MAE": [self.test_mae],
            'Before MAE': [self.finetune_before_mae],
            'After MAE': [self.finetune_after_mae],
            "Valid NRMSE": [self.valid_nrmse],
            "Test NRMSE": [self.test_nrmse],
            'Before NRMSE': [self.finetune_before_nrmse],
            'After NRMSE': [self.finetune_after_nrmse],
            'Before Stim MSE':[self.stim_before_mse],
            'After Stim MSE':[self.stim_after_mse],
            "Memo" : [self.config.memo]
        }

        # 创建一个DataFrame
        df = pd.DataFrame(data)

        file_path = self.config.result_path + "/" + 'Experiment_results.xlsx'


        if os.path.exists(file_path):
            df_before = pd.read_excel(file_path)
            df_before = pd.DataFrame(df_before)
            df = pd.concat([df_before, df], ignore_index=True)

        df.to_excel(file_path, index=False)

        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(file_path)










